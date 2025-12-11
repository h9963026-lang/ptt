#!/usr/bin/env python3
"""Count how many posts mention a keyword on every PTT board and export to Excel."""

from __future__ import annotations

import argparse
import json
import sys
import threading
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set
from urllib.parse import quote, quote_plus, urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook


BASE_URL = "https://www.ptt.cc"
DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
}


_thread_local = threading.local()


@dataclass
class WorkResult:
    board: str
    count: int
    pages: int
    per_keyword: Dict[str, int]
    errors: Dict[str, str] = field(default_factory=dict)
    error: Optional[str] = None


class BoardSearchError(RuntimeError):
    """Raised when we cannot obtain search results for a board."""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Search every PTT board for a keyword and export counts to Excel."
    )
    parser.add_argument(
        "--boards-file",
        default="看板名称/所有的看板名.json",
        help="Path to the board list (json / txt / xlsx). Default: %(default)s",
    )
    parser.add_argument(
        "--keywords",
        nargs="+",
        default=["诈骗", "詐騙", "詐欺"],
        help="Keywords used in the board search query (default: %(default)s).",
    )
    parser.add_argument(
        "--max-pages",
        type=int,
        default=0,
        help="Limit pages per board (0 = follow pagination until it stops).",
    )
    parser.add_argument(
        "--max-workers",
        type=int,
        default=5,
        help="Number of concurrent board searches (default: %(default)s).",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Only process the first N boards from the list (0 = all).",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.5,
        help="Minimum delay (seconds) between requests per thread (default: %(default)s).",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=12.0,
        help="HTTP timeout in seconds (default: %(default)s).",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=3,
        help="How many times to retry a failed request (default: %(default)s).",
    )
    parser.add_argument(
        "--sheet-name",
        default="诈骗统计",
        help="Excel sheet name (default: %(default)s).",
    )
    parser.add_argument(
        "--output",
        default="看板数据/诈骗帖子统计.xlsx",
        help="Output Excel path (default: %(default)s).",
    )
    return parser.parse_args()


def get_session() -> requests.Session:
    session = getattr(_thread_local, "session", None)
    if session is None:
        session = requests.Session()
        session.headers.update(DEFAULT_HEADERS)
        session.cookies.set("over18", "1", domain="ptt.cc")
        session.cookies.set("over18", "1", domain=".ptt.cc")
        _thread_local.session = session
        _thread_local.last_request = 0.0
    return session


def throttle(delay: float) -> None:
    if delay <= 0:
        return
    last = getattr(_thread_local, "last_request", 0.0)
    now = time.monotonic()
    remaining = delay - (now - last)
    if remaining > 0:
        time.sleep(remaining)
    _thread_local.last_request = time.monotonic()


def fetch_soup(url: str, timeout: float, delay: float, retries: int) -> BeautifulSoup:
    session = get_session()
    last_exc: Optional[Exception] = None
    for attempt in range(1, max(1, retries) + 1):
        throttle(delay)
        try:
            resp = session.get(url, timeout=timeout)
        except requests.RequestException as exc:
            last_exc = exc
        else:
            if resp.status_code == 200:
                resp.encoding = resp.apparent_encoding or "utf-8"
                return BeautifulSoup(resp.text, "html.parser")
            if resp.status_code == 404:
                raise BoardSearchError(f"{url} returned 404 (board missing?)")
            last_exc = RuntimeError(f"{url} -> HTTP {resp.status_code}")
        time.sleep(min(2.0, 0.7 * attempt))
    raise BoardSearchError(f"{url} failed after retries: {last_exc}")


def next_page_url(soup: BeautifulSoup) -> Optional[str]:
    for a in soup.select("div.btn-group-paging a"):
        text = (a.get_text() or "").strip()
        href = (a.get("href") or "").strip()
        if not href:
            continue
        if "上頁" in text:
            return urljoin(BASE_URL, href)
    return None


def collect_keyword_matches(
    board: str,
    keyword: str,
    max_pages: int,
    timeout: float,
    delay: float,
    retries: int,
) -> tuple[Set[str], int]:
    encoded_board = quote(board, safe="")
    base_url = f"{BASE_URL}/bbs/{encoded_board}/search?q={quote_plus(keyword)}"
    seen_pages: Set[str] = set()
    current_url: Optional[str] = base_url
    matches: Set[str] = set()
    visited = 0

    while current_url and (max_pages <= 0 or visited < max_pages):
        try:
            soup = fetch_soup(current_url, timeout=timeout, delay=delay, retries=retries)
        except BoardSearchError as exc:
            raise BoardSearchError(f"{keyword}: {exc}") from exc

        entries = soup.select("div.r-ent")
        for entry in entries:
            link = entry.select_one("a[href]")
            if not link:
                continue
            href = (link.get("href") or "").strip()
            if not href:
                continue
            matches.add(urljoin(BASE_URL, href))

        visited += 1

        if len(entries) == 0:
            break

        next_url = next_page_url(soup)
        if not next_url or next_url in seen_pages:
            break
        seen_pages.add(next_url)
        current_url = next_url

    return matches, visited


def count_keyword_on_board(
    board: str,
    keywords: Sequence[str],
    max_pages: int,
    timeout: float,
    delay: float,
    retries: int,
) -> WorkResult:
    all_matches: Set[str] = set()
    per_keyword: Dict[str, int] = {}
    keyword_errors: Dict[str, str] = {}
    pages = 0

    for keyword in keywords:
        try:
            matches, visited = collect_keyword_matches(
                board, keyword, max_pages, timeout, delay, retries
            )
        except BoardSearchError as exc:
            keyword_errors[keyword] = str(exc)
            continue
        per_keyword[keyword] = len(matches)
        all_matches.update(matches)
        pages += visited

    if not per_keyword:
        error_msg = "; ".join(keyword_errors.values()) or "No keywords succeeded."
        return WorkResult(
            board=board,
            count=0,
            pages=pages,
            per_keyword={},
            errors=keyword_errors,
            error=error_msg,
        )

    return WorkResult(
        board=board,
        count=len(all_matches),
        pages=pages,
        per_keyword=per_keyword,
        errors=keyword_errors,
        error=None,
    )


def load_board_names(path: Path) -> List[str]:
    if not path.exists():
        raise FileNotFoundError(f"Board list not found: {path}")

    suffix = path.suffix.lower()
    raw: List[str] = []

    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            raw = [str(v) for v in data.values()]
        elif isinstance(data, list):
            raw = [str(v) for v in data]
        else:
            raise ValueError("Unsupported JSON structure for board list.")
    elif suffix in {".txt", ".lst"}:
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                raw.append(line)
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            cell = row[0]
            if cell is None:
                continue
            value = str(cell).strip()
            if value:
                raw.append(value)
    else:
        raise ValueError(f"Unsupported board file format: {path.suffix}")

    cleaned: List[str] = []
    seen: set[str] = set()
    for value in raw:
        name = value.strip()
        if not name or "/" in name:
            continue
        if name in seen:
            continue
        seen.add(name)
        cleaned.append(name)
    return cleaned


def write_excel(results: Sequence[WorkResult], output: Path, sheet_name: str) -> None:
    """Merge the new stats into the Excel sheet instead of overwriting everything."""
    output.parent.mkdir(parents=True, exist_ok=True)

    if output.exists():
        wb = load_workbook(output)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(["board", "scam_posts"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["board", "scam_posts"])

    existing_rows = {}
    # Build index: board name -> cell row
    for row in ws.iter_rows(min_row=2):
        board_cell = row[0]
        if board_cell.value:
            existing_rows[str(board_cell.value)] = row

    for item in results:
        row = existing_rows.get(item.board)
        if row:
            row[1].value = item.count
        else:
            ws.append([item.board, item.count])
            existing_rows[item.board] = ws[ws.max_row]

    wb.save(output)


def process_boards(args: argparse.Namespace, boards: Sequence[str]) -> List[WorkResult]:
    from concurrent.futures import ThreadPoolExecutor, as_completed

    total = len(boards)
    results: Dict[str, WorkResult] = {}
    print(f"Processing {total} boards with keywords: {', '.join(args.keywords)} ...")

    with ThreadPoolExecutor(max_workers=args.max_workers) as executor:
        future_map = {
            executor.submit(
                count_keyword_on_board,
                board,
                args.keywords,
                args.max_pages,
                args.timeout,
                args.delay,
                args.retries,
            ): board
            for board in boards
        }

        for idx, future in enumerate(as_completed(future_map), start=1):
            board = future_map[future]
            try:
                result = future.result()
            except Exception as exc:  # pragma: no cover - defensive
                result = WorkResult(
                    board=board,
                    count=0,
                    pages=0,
                    per_keyword={},
                    errors={},
                    error=str(exc),
                )
            results[board] = result
            status = "OK" if not result.error else f"ERR: {result.error}"
            per_kw = ", ".join(f"{k}:{v}" for k, v in result.per_keyword.items())
            if per_kw:
                status += f" | {per_kw}"
            if result.errors:
                errs = ", ".join(f"{k}:{v}" for k, v in result.errors.items())
                status += f" | keyword errors -> {errs}"
            print(f"[{idx}/{total}] {board}: {result.count} matches ({result.pages} pages) {status}")

    ordered = [results[board] for board in boards if board in results]
    return ordered


def main() -> None:
    args = parse_args()
    try:
        boards = load_board_names(Path(args.boards_file))
    except Exception as exc:
        print(f"Failed to load board list: {exc}", file=sys.stderr)
        sys.exit(1)

    if args.limit > 0:
        boards = boards[: args.limit]

    if not boards:
        print("Board list is empty.", file=sys.stderr)
        sys.exit(1)

    results = process_boards(args, boards)
    write_excel(results, Path(args.output), args.sheet_name)

    errors = [item for item in results if item.error]
    print(
        f"Done. {len(results)} boards processed; "
        f"{sum(item.count for item in results)} matching posts counted; "
        f"{len(errors)} boards had errors."
    )
    if errors:
        print("Boards with errors:")
        for item in errors:
            print(f"  - {item.board}: {item.error}")


if __name__ == "__main__":
    main()
