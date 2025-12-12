#!/usr/bin/env python3
import argparse
import gzip
import json
import math
import os
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, List, Dict, Optional, Set, Tuple
from urllib.parse import quote_plus

import pandas as pd  # 用于读取Excel
import requests
from bs4 import BeautifulSoup


# === 配置信息（已根据需求编码）===
# Excel看板路径
EXCEL_PATH = r"F:\SEMESTER A\5507\5507PTT\工作簿1.xlsx"
# 存储路径
STORAGE_DIR = r"F:\SEMESTER A\5507\5507PTT\看板数据"
# 输出文件名称
JSON_OUTPUT = os.path.join(STORAGE_DIR, "ptt_posts.json")
EXCEL_OUTPUT = os.path.join(STORAGE_DIR, "ptt_posts.xlsx")
RAW_JSONL = os.path.join(STORAGE_DIR, "ptt_raw.jsonl")
# 关键词（默认诈骗相关）
KEYWORDS = ["诈骗", "詐騙", "詐欺"]
# PTT网页端域名
BASE_URL = "https://www.ptt.cc"
# 其他配置
EXPORT_COLUMNS = ["board", "title", "url", "article_id", "saved_html", "saved_at"]
SAVED_HTML_COL = EXPORT_COLUMNS.index("saved_html")


# === 工具函数 ===
def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def safe_filename(name: str) -> str:
    name = name.strip().replace("/", "-")
    name = re.sub(r"[\\\\:*?\"<>|]", "_", name)
    name = re.sub(r"\s+", " ", name)
    return name[:200]


def extract_article_id(url: str) -> str:
    m = re.search(r"/M\.([\w\.]+)\.html$", url)
    if m:
        return m.group(1)
    return re.sub(r"\W+", "_", url)


def sanitize_html(html: str) -> str:
    if "ask/over18" not in html:
        return html
    try:
        soup = BeautifulSoup(html, "html.parser")
        for script in soup.find_all("script"):
            content = script.string or script.get_text() or ""
            if "ask/over18" in content:
                script.decompose()
        return str(soup)
    except Exception:
        pass
    pattern = re.compile(r"<script[^>]*>.*?ask/over18.*?</script>", re.DOTALL | re.IGNORECASE)
    return pattern.sub("", html)


def load_existing_raw_urls(path: Optional[str]) -> Set[str]:
    urls: Set[str] = set()
    if not path or not os.path.exists(path):
        return urls
    opener = gzip.open if path.endswith(".gz") else open
    try:
        with opener(path, "rt", encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    data = json.loads(line)
                except json.JSONDecodeError:
                    continue
                url = data.get("url")
                if isinstance(url, str) and url:
                    urls.add(url)
    except Exception:
        return urls
    return urls


def append_raw_record(path: Optional[str], record: Dict[str, Any], compress: bool = False) -> None:
    if not path:
        return
    use_gzip = compress or path.endswith(".gz")
    opener = gzip.open if use_gzip else open
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    with opener(path, "at", encoding="utf-8") as f:
        line = json.dumps(record, ensure_ascii=False)
        f.write(line + "\n")


# === 从Excel读取看板名称 ===
def get_board_names_from_excel() -> List[str]:
    """读取工作簿1中的看板名称"""
    try:
        df = pd.read_excel(EXCEL_PATH)
        if "看板名称" not in df.columns:
            print("Excel中未找到“看板名称”列，请检查表头！")
            return []
        # 提取非空且去重的看板名称
        board_names = df["看板名称"].dropna().unique().tolist()
        print(f"从Excel成功读取 {len(board_names)} 个看板名称")
        return [str(b) for b in board_names]
    except Exception as e:
        print(f"读取Excel失败：{str(e)}")
        return []


# === PTT爬虫类 ===
class PTTCrawler:
    def __init__(self, delay: float = 1.0, timeout: float = 15.0):
        self.session = requests.Session()
        self.session.cookies.set("over18", "1", domain="ptt.cc")  # 自动同意18+
        self.delay = max(0.0, delay)
        self.timeout = timeout
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
            "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
        }

    def _sleep(self):
        if self.delay > 0:
            time.sleep(self.delay)

    def _get(self, url: str) -> requests.Response:
        self._sleep()
        resp = self.session.get(url, headers=self.headers, timeout=self.timeout)
        resp.encoding = resp.apparent_encoding or "utf-8"
        return resp

    def _full(self, path: str) -> str:
        if path.startswith("http://") or path.startswith("https://"):
            return path
        return BASE_URL + path

    def iter_board_pages(self, board: str, max_pages: int) -> Iterable[BeautifulSoup]:
        pages_fetched = 0
        url = f"{BASE_URL}/bbs/{board}/index.html"
        while pages_fetched < max_pages:
            try:
                r = self._get(url)
            except Exception:
                break
            if r.status_code != 200:
                break
            soup = BeautifulSoup(r.text, "html.parser")
            yield soup
            pages_fetched += 1
            # 查找上一页链接
            prev = None
            for a in soup.select("div.btn-group-paging a"):
                txt = a.get_text(strip=True)
                if "上頁" in txt:
                    prev = a
                    break
            if not prev or not prev.get("href"):
                break
            url = self._full(prev.get("href"))

    def iter_search_pages(
        self,
        board: str,
        query: str,
        max_pages: int,
        page_sleep: float = 0.0,
    ) -> Iterable[BeautifulSoup]:
        max_pages = max(0, int(max_pages or 0))
        if max_pages <= 0:
            return
        pages_fetched = 0
        url = f"{BASE_URL}/bbs/{board}/search?q={quote_plus(query)}"
        while pages_fetched < max_pages:
            try:
                r = self._get(url)
            except Exception:
                break
            if r.status_code != 200:
                break
            soup = BeautifulSoup(r.text, "html.parser")
            yield soup
            pages_fetched += 1
            next_href = None
            for a in soup.select("div.btn-group-paging a"):
                txt = a.get_text(strip=True)
                if "上頁" in txt and a.get("href"):
                    next_href = a.get("href")
                    break
            if not next_href:
                break
            url = self._full(next_href)
            if page_sleep > 0:
                time.sleep(page_sleep)

    def extract_posts_from_index(self, soup: BeautifulSoup) -> List[Dict[str, str]]:
        posts: List[Dict[str, str]] = []
        for ent in soup.select("div.r-ent"):
            a = ent.select_one("div.title a")
            if not a:
                continue
            title = a.get_text(strip=True)
            href = a.get("href")
            if not href:
                continue
            url = self._full(href)
            posts.append({"title": title, "url": url})
        return posts

    @staticmethod
    def match_keywords(text: str, keywords: List[str]) -> bool:
        t = text or ""
        return any(k in t for k in keywords)

    def fetch_post_html(self, url: str) -> Optional[str]:
        try:
            r = self._get(url)
            if r.status_code != 200:
                return None
            return r.text
        except Exception:
            return None


# === 核心爬取逻辑 ===
def _keyword_match_scope(
    title: str,
    raw_html: Optional[str],
    keywords: List[str],
    scope: str,
) -> Tuple[bool, Optional[str]]:
    if not keywords:
        return True, None
    matched_keyword: Optional[str] = None
    scope = scope or "title"
    title = title or ""
    if scope in {"title", "both"}:
        for kw in keywords:
            if kw in title:
                return True, kw
    if scope in {"body", "both"} and raw_html:
        for kw in keywords:
            if kw in raw_html:
                return True, kw
    return False, matched_keyword


def crawl_boards(
    boards: List[str],
    keywords: List[str],
    out_dir: str,
    max_pages: int,
    delay: float,
    timeout: float,
    skip_existing: bool,
    use_search: bool,
    search_pages: int,
    page_sleep: float,
    download_sleep: float,
    max_downloads_per_batch: int,
    fallback_index_pages: int,
    raw_jsonl_path: Optional[str],
    write_raw_jsonl: bool,
    raw_jsonl_compress: bool,
    save_html: bool,
    existing_raw_urls: Optional[Set[str]],
    force_raw_when_skip: bool,
    full_index_scan: bool,
    max_index_pages: int,
    match_scope: str,
) -> List[Dict[str, str]]:
    if save_html:
        os.makedirs(out_dir, exist_ok=True)
    crawler = PTTCrawler(delay=delay, timeout=timeout)

    normalized_keywords = [k for k in dict.fromkeys(keywords) if k]
    max_pages = max(0, int(max_pages or 0))
    search_pages = max(0, int(search_pages or 0))
    fallback_index_pages = max(0, int(fallback_index_pages or 0))
    max_index_pages = max(0, int(max_index_pages or 0))
    max_downloads_per_batch = max(0, int(max_downloads_per_batch or 0))
    page_sleep = max(0.0, page_sleep)
    download_sleep = max(0.0, download_sleep)
    match_scope = (match_scope or "title").lower()

    all_records: List[Dict[str, str]] = []
    raw_urls_seen: Set[str] = existing_raw_urls if existing_raw_urls is not None else set()

    for board in boards:
        print(f"[+] 正在爬取看板: {board}")
        board_dir = os.path.join(out_dir, board)
        if save_html:
            os.makedirs(board_dir, exist_ok=True)

        seen_urls: Set[str] = set()
        candidate_posts: List[Dict[str, str]] = []

        # 1) 使用搜索功能查找关键词
        if use_search and not full_index_scan:
            if search_pages <= 0:
                print("  - 搜索已启用，但搜索页数为0，跳过搜索")
            else:
                for keyword in normalized_keywords or [""]:
                    page_idx = 0
                    before_count = len(candidate_posts)
                    for soup in crawler.iter_search_pages(board, keyword, search_pages, page_sleep):
                        page_idx += 1
                        posts = crawler.extract_posts_from_index(soup)
                        collected = 0
                        for p in posts:
                            url = p.get("url")
                            if not url or url in seen_urls:
                                continue
                            title = p.get("title", "")
                            if keyword and match_scope == "title" and not PTTCrawler.match_keywords(title, [keyword]):
                                continue
                            seen_urls.add(url)
                            p["keyword"] = keyword
                            candidate_posts.append(p)
                            collected += 1
                        print(
                            f"  - 关键词 '{keyword}' 第 {page_idx} 页: "
                            f"共 {len(posts)} 篇，新增 {collected} 篇，累计 {len(candidate_posts)} 篇"
                        )
                    print(
                        f"  - 关键词 '{keyword}' 共收集 {len(candidate_posts) - before_count} 篇候选帖子"
                    )

        # 2) 若搜索无结果，遍历看板索引页
        need_index_scan = full_index_scan or (not candidate_posts and fallback_index_pages > 0)
        if need_index_scan:
            limit = max_index_pages if full_index_scan else fallback_index_pages
            if limit <= 0:
                limit = 10**9
            source = "索引页" if full_index_scan else "备用索引页"
            page_i = 0
            for soup in crawler.iter_board_pages(board, max_pages=limit):
                page_i += 1
                posts = crawler.extract_posts_from_index(soup)
                collected = 0
                for p in posts:
                    url = p.get("url")
                    if not url or url in seen_urls:
                        continue
                    title = p.get("title", "")
                    if normalized_keywords and match_scope == "title" and not PTTCrawler.match_keywords(title, normalized_keywords):
                        continue
                    seen_urls.add(url)
                    candidate_posts.append(p)
                    collected += 1
                print(
                    f"  - {source} 第 {page_i} 页: 共 {len(posts)} 篇，新增 {collected} 篇，"
                    f"累计 {len(candidate_posts)} 篇"
                )

        print(f"  - 看板 {board} 候选帖子总数: {len(candidate_posts)}")

        saved = 0
        skipped_existing = 0
        failed = 0

        for p in candidate_posts:
            if max_downloads_per_batch and saved >= max_downloads_per_batch:
                print(f"  - 已达到每批下载上限 ({max_downloads_per_batch})，暂停当前看板")
                break

            url = p.get("url")
            if not url:
                continue
            already_stored = write_raw_jsonl and url in raw_urls_seen

            aid = extract_article_id(url) or "article"
            title = safe_filename(p.get("title", "untitled"))
            fname = f"{aid} - {title}.html"
            fpath = os.path.join(board_dir, fname)
            saved_path = ""

            if already_stored and not force_raw_when_skip:
                skipped_existing += 1
                if skip_existing:
                    record = {
                        "board": board,
                        "title": p.get("title", ""),
                        "url": url,
                        "article_id": aid,
                        "saved_html": saved_path,
                        "saved_at": utc_now_iso(),
                    }
                    keyword = p.get("keyword")
                    if keyword:
                        record["keyword"] = keyword
                    all_records.append(record)
                continue

            raw_html = crawler.fetch_post_html(url)
            if download_sleep > 0:
                time.sleep(download_sleep)
            if not raw_html:
                failed += 1
                continue

            timestamp = utc_now_iso()

            should_save, matched_keyword = _keyword_match_scope(
                p.get("title", ""),
                raw_html,
                normalized_keywords,
                match_scope,
            )
            if not should_save:
                continue

            if save_html:
                try:
                    with open(fpath, "w", encoding="utf-8") as f:
                        f.write(raw_html)
                    saved_path = os.path.abspath(fpath)
                except Exception:
                    failed += 1
                    continue

            if write_raw_jsonl:
                try:
                    raw_record = {
                        "url": url,
                        "created_at": timestamp,
                        "html": raw_html,
                        "source": board,
                    }
                    append_raw_record(raw_jsonl_path, raw_record, raw_jsonl_compress)
                    raw_urls_seen.add(url)
                except Exception:
                    failed += 1
                    continue

            record = {
                "board": board,
                "title": p.get("title", ""),
                "url": url,
                "article_id": aid,
                "saved_html": saved_path,
                "saved_at": timestamp,
            }
            keyword = matched_keyword or p.get("keyword")
            if keyword:
                record["keyword"] = keyword

            saved += 1
            all_records.append(record)

        print(
            f"  - 看板 {board} 处理完成: 新增 {saved} 篇，跳过 {skipped_existing} 篇，失败 {failed} 篇"
        )

    return all_records


# === 写入输出文件 ===
def _dedup_preserve(items: Iterable[str]) -> List[str]:
    seen: Set[str] = set()
    result: List[str] = []
    for item in items:
        if not item:
            continue
        lowered = item.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        result.append(item)
    return result


def write_json(records: List[Dict[str, str]], path: str) -> List[Dict[str, str]]:
    if not path:
        return records

    existing: List[Dict[str, str]] = []
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            if isinstance(loaded, list):
                existing = loaded
        except Exception:
            existing = []

    merged: List[Dict[str, str]] = []
    index: Dict[Tuple[Optional[str], Optional[str]], int] = {}

    for item in existing:
        key = (item.get("board"), item.get("article_id"))
        index[key] = len(merged)
        merged.append(item)

    for rec in records:
        key = (rec.get("board"), rec.get("article_id"))
        if key in index:
            merged[index[key]] = rec
        else:
            index[key] = len(merged)
            merged.append(rec)

    dirpath = os.path.dirname(path)
    if dirpath:
        os.makedirs(dirpath, exist_ok=True)

    with open(path, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)

    return merged


def write_excel(records: List[Dict[str, str]], path: str) -> None:
    if not records or not path:
        return

    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise SystemExit("请安装openpyxl以支持Excel输出：pip install openpyxl") from exc

    existing_rows: Dict[Tuple[Optional[str], Optional[str]], Tuple] = {}

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            key = (row[0].value, row[3].value)
            existing_rows[key] = row
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "PTT帖子数据"
        ws.append(EXPORT_COLUMNS)

    ws = wb.active

    for record in records:
        key = (record.get("board"), record.get("article_id"))
        values = [record.get(col, "") for col in EXPORT_COLUMNS]
        if key in existing_rows:
            row_cells = existing_rows[key]
            for idx, cell in enumerate(row_cells):
                if idx < len(values):
                    cell.value = values[idx]
        else:
            ws.append(values)

    wb.save(path)


# === 主函数 ===
def main():
    # 创建存储目录
    os.makedirs(STORAGE_DIR, exist_ok=True)
    
    # 从Excel读取看板
    boards = get_board_names_from_excel()
    if not boards:
        print("未读取到任何看板名称，程序退出")
        return

    # 配置参数
    write_raw_jsonl = True
    raw_urls_seen = load_existing_raw_urls(RAW_JSONL)
    print(f"已加载 {len(raw_urls_seen)} 个已爬取的URL（用于去重）")

    batch_size = 10
    total_boards = len(boards)
    total_batches = math.ceil(total_boards / batch_size)

    keyword_groups = [KEYWORDS]
    all_new_records: List[Dict[str, str]] = []

    for keyword_index, keyword_list in enumerate(keyword_groups, start=1):
        keyword_label = "、".join(keyword_list)
        print(f"\n===== 关键词组 {keyword_index}/{len(keyword_groups)}: {keyword_label} =====")

        for batch_index in range(total_batches):
            start = batch_index * batch_size
            batch_boards = boards[start : start + batch_size]
            batch_label = batch_index + 1
            print(f"\n----- 批次 {batch_label}/{total_batches}: 看板列表 {batch_boards} -----")

            new_records = crawl_boards(
                boards=batch_boards,
                keywords=keyword_list,
                out_dir=os.path.join(STORAGE_DIR, "ptt_html"),  # HTML子目录
                max_pages=10,
                delay=1.2,
                timeout=15.0,
                skip_existing=True,
                use_search=True,
                search_pages=500,
                page_sleep=1.0,
                download_sleep=0.5,
                max_downloads_per_batch=0,
                fallback_index_pages=0,
                raw_jsonl_path=RAW_JSONL,
                write_raw_jsonl=write_raw_jsonl,
                raw_jsonl_compress=False,
                save_html=True,
                existing_raw_urls=raw_urls_seen,
                force_raw_when_skip=False,
                full_index_scan=False,
                max_index_pages=0,
                match_scope="title",
            )

            all_new_records.extend(new_records)
            print(f"批次 {batch_label} 完成，累计新增 {len(all_new_records)} 条数据")

    # 写入结果文件
    merged_records = write_json(all_new_records, JSON_OUTPUT)
    write_excel(merged_records, EXCEL_OUTPUT)
    print(f"\n所有爬取完成！结果已保存至：")
    print(f"  - JSON元数据：{JSON_OUTPUT}")
    print(f"  - Excel表格：{EXCEL_OUTPUT}")
    print(f"  - 原始HTML文件：{os.path.join(STORAGE_DIR, 'ptt_html')}")
    print(f"  - 原始数据JSONL：{RAW_JSONL}")


if __name__ == "__main__":
    main()

